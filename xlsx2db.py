"""
# XLSX2DB
XLSX 파일을 DB 테이블에 입력합니다.
MSSQL, MySQL, MariaDB를 지원합니다.

## Contributor
- 손남규(snk@fasoo.com)
"""

# 내부 모듈
import datetime
import configparser
import pprint   # 개발용
pp = pprint.PrettyPrinter(indent=4, sort_dicts=False)

# 외부 모듈
import openpyxl
import pymssql
import pymysql

class XLSX2DB:
    """
    # XSLX2DB
    XLSX 파일을 읽어 DB에 INSERT 합니다.
    """
    __debug: bool = False                   # 개발용
    dbms_name: str                          # DBMS 종류
    dbms_server: str                        # DBMS 서버 주소
    dbms_username: str                      # DBMS 사용자 이름
    dbms_password: str                      # DBMS 사용자 암호
    database_name: str                      # 데이터베이스 이름


    def __init__(self, section: str = "DEFAULT") -> None:
        config = configparser.ConfigParser()
        config.read("xlsx2db.ini")
        self.dbms_name = config[section]["dbms_name"]
        self.dbms_server = config[section]["dbms_server"]
        self.dbms_username = config[section]["dbms_username"]
        self.dbms_password = config[section]["dbms_password"]
        self.database_name = config[section]["database_name"]


    def convert_xlsx2db(self, path: str, table: str) -> bool:
        """
        # convert_xlsx2db()
        XLSX 파일을 읽어, 1행에 있는 컬럼명을 기준으로 DB에 데이터를 삽입합니다.

        ## Args:
        1. path: str = "example.xlsx"

        ## Returns:
        output: bool = True # 정상실행일 경우 True
        """
        output: bool = True

        # Part I. Loading Excel File
        print("{dt}: Start of loading the file...".format(dt = datetime.datetime.now().strftime("%Y-%m-%d T %H:%M:%S")))
        column_names, records = self.load_xlsx(path)
        print("{dt}: End of loading the file...".format(dt = datetime.datetime.now().strftime("%Y-%m-%d T %H:%M:%S")))

        # Part II. Create Query
        print("{dt}: Start of creating a query...".format(dt = datetime.datetime.now().strftime("%Y-%m-%d T %H:%M:%S")))
        query = self.create_query(table, column_names)
        print("{dt}: End of creating a query...".format(dt = datetime.datetime.now().strftime("%Y-%m-%d T %H:%M:%S")))

        # Part III. Execute Query
        print("{dt}: Start of executing the query...".format(dt = datetime.datetime.now().strftime("%Y-%m-%d T %H:%M:%S")))
        output = self.execute_query(query, records)
        print("{dt}: End of creating the query...".format(dt = datetime.datetime.now().strftime("%Y-%m-%d T %H:%M:%S")))
        
        return output


    # Part I. Loading Excel XLSX File
    def load_xlsx(self, path: str) -> list:
        """
        # load_xlsx()
        XLSX 파일을 읽어, 1행에 있는 컬럼명을 기준으로 DB에 삽입하기 위한 리스트를 생성합니다.

        ## Args:
        1. path: str = "example.xlsx"

        ## Returns:
        output: tuple = (column_names, records)
        """
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active

        records: list = []
        i: int = 0
        for row in worksheet.iter_rows(values_only=True):
            if i == 0:
                # Names of column(컬럼명)
                column_names = list(row)
            else:
                # Values of record(레코드값)
                # records.append(dict(zip(column_names, list(row))))
                records.append(list(row))
            i += 1

        if self.__debug:
            print("column_names = ")
            pp.pprint(column_names)
            print("output = ")
            pp.pprint(records)

        return (column_names, records)


    # Part II. Create Query
    def create_query(self, table: str, column_names: list) -> str:
        """
        # create_query()
        쿼리를 생성합니다.

        ## Args:
        1. table: str = "example"
        2. column_names: list = ["column1", ...]

        ## Returns:
        output: str = "INSERT INTO example (column1, ...) VALUES ('value1', ...); ..."
        """
        columns = ", ".join(column_names)
        values_placeholder = ", ".join(["%s"]*len([column_name for column_name in column_names if column_name is not None]))
        output = f"""INSERT INTO {table} ({columns}) VALUES ({values_placeholder})"""

        if self.__debug:
            print("query = ")
            print(output)

        return output
    

    # Part III. Execute Query
    def execute_query(self, query: str, records: list) -> bool:
        """
        # execute_query()
        쿼리를 여러번 실행합니다.

        ## Args:
        1. query: str = "INSERT INTO example (column1, ...) VALUES ('value1', ...); ..."
        2. records: list = [("value1", ...), ...]

        ## Returns:
        output: bool = True # 정상실행일 경우 True
        """
        output: bool = True

        try:
            if self.dbms_name.lower() == "mssql":
                connection = pymssql.connect(server = self.dbms_server, user = self.dbms_username, password = self.dbms_password, database = self.database_name)
            elif self.dbms_name.lower() == "mysql" or self.dbms_name.lower() == "mariadb":
                connection = pymysql.connect(host = self.dbms_server, user = self.dbms_username, password = self.dbms_password, db = self.database_name, charset = "utf8")
            cursor = connection.cursor()
            cursor.executemany(query, records)
            connection.commit()
            connection.close()
        except Exception as e:
            if self.__debug:
                print("Database Error: ", e)
            output = False

        return output


    # Methods for manual run
    def convert_xlsx2dictlist(self, path: str) -> list:
        """
        # convert_xlsx2list()
        XLSX 파일을 읽어, 1행에 있는 컬럼명을 기준으로 DB에 삽입하기 위한 리스트를 생성합니다.

        ## Args:
        1. path: str = "example.xlsx"                       # 엑셀 파일 경로.

        ## Returns:
        output: list = [{"column1": "value1", ...}, ...]    # DB에 삽입하기 위한 Dictionary로 이루어진 리스트. 리스트 내 개별 원소 값이 레코드 하나에 대응함.
        """
        output: list = []

        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active

        print("{dt}: Start of loading the file...".format(dt = datetime.datetime.now().strftime("%Y-%m-%d T %H:%M:%S")))
        i: int = 0
        for row in worksheet.iter_rows(values_only=True):
            if i == 0:
                # Names of column(컬럼명)
                column_names = list(row)
            else:
                # Values of record(레코드값)
                output.append({column_names[i]: list(row)[i] for i in range(len(column_names)) if column_names[i] is not None})
            i += 1
        print("{dt}: End of loading the file...".format(dt = datetime.datetime.now().strftime("%Y-%m-%d T %H:%M:%S")))

        return output


    def convert_dictlist2sql(self, table: str, data: list) -> str:
        """
        # convert_dictlist2sql()

        ## Args:
        1. data: list = [{"column1": "value1", ...}, ...]   # DB에 삽입하기 위한 Dictionary로 이루어진 리스트. 리스트 내 개별 원소 값이 레코드 하나에 대응함.

        ## Returns:
        output: str = "INSERT INTO example (column1, ...) VALUES ('value1', ...); ..."
        """
        output: str = ""

        for row in data:
            columns = ', '.join(list(row.keys()))
            values = ','.join(["'" + v + "'" for v in list(row.values())])

            pp.pprint(values)

            output += f"INSERT INTO {table} ({columns}) VALUES ({values});\n"

        return output


    def execute(self, query: str) -> bool:
        """
        # execute()
        쿼리를 실행합니다.

        ## Args:
        1. query: str = "INSERT INTO example (column1, ...) VALUES ('value1', ...); ..."

        ## Returns:
        output: bool = True # 정상실행일 경우 True
        """
        output: bool = True

        try:
            if self.dbms_name.lower() == "mssql":
                connection = pymssql.connect(server = self.dbms_server, user = self.dbms_username, password = self.dbms_password, database = self.database_name, charset = "utf8")
            elif self.dbms_name.lower() == "mysql" or self.dbms_name.lower() == "mariadb":
                connection = pymysql.connect(host = self.dbms_server, user = self.dbms_username, password = self.dbms_password, db = self.database_name, charset = "utf8")
            cursor = connection.cursor()
            cursor.execute(query)
            connection.commit()
            connection.close()
        except:
            output = False

        return output


if __name__ == "__main__":
    xlsx2db = XLSX2DB(section="DEFAULT") # section은 설정 파일(xlsx2db.ini)의 섹션명
    output: bool = xlsx2db.convert_xlsx2db(path="example.xlsx", table="example") # path는 엑셀 파일의 경로, table은 DB 테이블의 이름
    print(output) # True인 경우 정상 종료